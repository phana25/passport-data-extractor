import os
import string as st
import ssl
import shutil
import sys
from dateutil import parser
import matplotlib.image as mpimg
import cv2
from passporteye import read_mrz
import json
import easyocr
import pytesseract
import numpy as np
import openpyxl
import warnings
import tempfile
import re
import certifi

warnings.filterwarnings('ignore')


def _configure_ssl_certificates():
    """
    Ensure HTTPS downloads (e.g. EasyOCR model files) use a trusted CA bundle.
    This is especially important on some Windows Python/venv setups.
    """
    cafile = certifi.where()
    os.environ.setdefault("SSL_CERT_FILE", cafile)
    os.environ.setdefault("REQUESTS_CA_BUNDLE", cafile)
    ssl._create_default_https_context = lambda: ssl.create_default_context(cafile=cafile)

class PassportDataExtractor:
    def __init__(self, country_codes_file, gpu=True):
        _configure_ssl_certificates()
        self._tesseract_available = self._configure_tesseract()
        
        # EasyOCR reader initialization — fall back to CPU if GPU initialization hangs or fails.
        # This is especially common on macOS where CUDA is not available.
        # Use a timeout or catch-all to prevent hangs on Windows with bad GPU drivers.
        effective_gpu = False
        if sys.platform == "win32" and gpu:
            # On Windows, try GPU but fall back to CPU immediately if it fails
            try:
                self.reader = easyocr.Reader(lang_list=['en'], gpu=True)
                effective_gpu = True
            except Exception:
                self.reader = easyocr.Reader(lang_list=['en'], gpu=False)
        else:
            self.reader = easyocr.Reader(lang_list=['en'], gpu=False)
            
        with open(country_codes_file) as f:
            self.country_codes = json.load(f)


    def _configure_tesseract(self):
        """
        Try to locate Tesseract executable across platforms.
        Returns True when available; False otherwise.
        """
        configured = getattr(pytesseract.pytesseract, "tesseract_cmd", "") or ""
        if configured and os.path.exists(configured):
            return True

        # Try system PATH first.
        exe = shutil.which("tesseract")
        if exe:
            pytesseract.pytesseract.tesseract_cmd = exe
            return True

        # Candidates for bundled binaries (PyInstaller) and common system locations.
        candidates = []
        is_windows = sys.platform == "win32"
        binary_name = "tesseract.exe" if is_windows else "tesseract"
        
        if getattr(sys, "frozen", False):
            # 1) Next to executable (one-folder)
            exe_dir = os.path.dirname(sys.executable)
            candidates.append(os.path.join(exe_dir, "tesseract", binary_name))
            # 2) Temp extraction dir (one-file)
            meipass = getattr(sys, "_MEIPASS", "")
            if meipass:
                candidates.append(os.path.join(meipass, "tesseract", binary_name))
        
        # 3) Project local vendor folder (dev mode)
        here = os.path.dirname(os.path.abspath(__file__))
        candidates.append(os.path.join(here, "vendor", "tesseract", binary_name))

        if not is_windows:
            # 4) Common macOS Homebrew paths (Apple Silicon & Intel)
            candidates.extend([
                "/opt/homebrew/bin/tesseract",
                "/usr/local/bin/tesseract",
                "/usr/bin/tesseract"
            ])
        else:
            # 5) Common Windows install locations
            candidates.extend([
                r"C:\Program Files\Tesseract-OCR\tesseract.exe",
                r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
            ])

        for candidate in candidates:
            if os.path.exists(candidate):
                pytesseract.pytesseract.tesseract_cmd = candidate
                # Configure TESSDATA_PREFIX for bundled distribution
                tessdata_dir = os.path.join(os.path.dirname(candidate), "tessdata")
                if os.path.isdir(tessdata_dir):
                    os.environ.setdefault("TESSDATA_PREFIX", tessdata_dir)
                return True

        return False


    # Common OCR misreads of month abbreviations
    _MONTH_FIXES = {
        'JAM': 'JAN', 'JAN': 'JAN',
        'FEB': 'FEB',
        'MAR': 'MAR',
        'APR': 'APR',
        'MAY': 'MAY',
        'JUM': 'JUN', 'JUN': 'JUN',
        'JUL': 'JUL',
        'AUG': 'AUG',
        'SEP': 'SEP',
        'OCT': 'OCT',
        'NOV': 'NOV',
        'DEC': 'DEC',
    }

    def _fix_month_typos(self, text):
        """
        Replace common OCR month misreads (JAM→JAN, JUM→JUN, etc.) and
        strip Chinese/mixed date noise like '10月/OCT' → 'OCT', '10A/OCT' → 'OCT'.
        """
        # Remove Chinese/mixed month prefix patterns: digits + chars + slash before month
        # e.g. "10A/OCT" → "OCT", "10月/OCT" → "OCT", "romvoct" → "OCT"
        text = re.sub(
            r'\d+[^\s/]*/\s*(?=(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC))',
            '', text, flags=re.IGNORECASE
        )
        # Fix garbled months embedded in junk (e.g. "romvoct" → "OCT", "jaN" → "JAN")
        text = re.sub(
            r'\b\w*?(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\w*\b',
            lambda m: m.group(1).upper(),
            text, flags=re.IGNORECASE
        )

        def _replace(m):
            return self._MONTH_FIXES.get(m.group(0).upper(), m.group(0))
        return re.sub(r'\b(JAM|JUM|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)\b',
                      _replace, text, flags=re.IGNORECASE)

    def _tesseract_lines(self, img):
        """Run Tesseract on a cv2 image and return non-empty lines."""
        if not self._tesseract_available:
            return []
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        gray = cv2.GaussianBlur(gray, (3, 3), 0)
        _, thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        try:
            text = pytesseract.image_to_string(thresh, config='--psm 6 --oem 3')
        except pytesseract.pytesseract.TesseractNotFoundError:
            self._tesseract_available = False
            return []
        lines = [self._fix_month_typos(l.strip()) for l in text.splitlines() if l.strip()]
        return lines

    def _easyocr_lines(self, img):
        """Run EasyOCR only, with month-typo fixes applied."""
        return [self._fix_month_typos(l) for l in self.reader.readtext(img, detail=0)]

    def _dual_ocr_lines(self, img):
        """Merge EasyOCR + Tesseract — keep all lines from both engines."""
        return self._easyocr_lines(img) + self._tesseract_lines(img)

    def parse_birth_date(self, date_string):
        try:
            # MRZ DOB is YYMMDD; dateutil may pick a future century for some values.
            # Keep it in a realistic passport holder range: if parsed year is in the future,
            # roll it back by 100 years (e.g. 2088 -> 1988). Do not force 2001 -> 1901.
            import datetime as _dt
            date = parser.parse(date_string, yearfirst=True).date()
            if date.year > _dt.date.today().year:
                date = date.replace(year=date.year - 100)
            return date.strftime('%d/%m/%Y')
        except ValueError:
            return None

    def parse_date(self, date_string, dayfirst=False):
        try:
            date = parser.parse(date_string, yearfirst=not dayfirst, dayfirst=dayfirst).date()
            return date.strftime('%d/%m/%Y')
        except (ValueError, TypeError):
            return None

    def _parse_ocr_date(self, date_string):
        """Parse date from OCR using dayfirst. Strips noise like dots in day (e.g. '13. JAN 2022')."""
        import datetime as _dt
        # 1. Fuzzy fixes for common OCR misreads in dates (S -> 5, I/l -> 1, O -> 0)
        # We only do this locally to avoid corrupting other fields.
        cleaned = date_string.strip().upper()
        
        # If the string looks like a date but has 'S', 'I', 'O', 'l', 'B' where digits should be
        # e.g. "l3 JAN 2022" or "13 JAN 2O22"
        if re.match(r'^[L|I|S|O|B|0-9].*', cleaned):
            # Day part starting
            cleaned = re.sub(r'^([L|I])(\d)\b', r'1\2', cleaned)
            cleaned = re.sub(r'^(\d)([L|I])\b', r'\g<1>1', cleaned)
            cleaned = re.sub(r'^(S)(\d)\b', r'5\2', cleaned)
            cleaned = re.sub(r'^(\d)(S)\b', r'\g<1>5', cleaned)
            cleaned = re.sub(r'^(O)(\d)\b', r'0\2', cleaned)
            cleaned = re.sub(r'^(\d)(O)\b', r'\g<1>0', cleaned)
            
            # Year part (at the end)
            cleaned = re.sub(r'\b(2[O|0]2[L|I|1])\b', r'2021', cleaned)
            cleaned = re.sub(r'\b(2[O|0]2[O|0])\b', r'2020', cleaned)
            cleaned = re.sub(r'\b2[O|0][L|I|1](\d)\b', r'201\1', cleaned)
            # Generic catch remaining O in 4-digit years (like 2O28)
            # but only securely if it already looks like a year
            if "2O" in cleaned:
                cleaned = cleaned.replace("2O", "20")

        # After fuzzy fixes, strip out garbage like 'SA/' in '05 SA/MAR 2028'
        # or '03.AUG 1991' -> '03 AUG 1991'
        months = 'JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC'
        m = re.search(r'([0-9]{1,2})[.,\sA-Z/]+(' + months + r')[A-Z]*[.,\s/]+([0-9]{2,4})', cleaned)
        if m:
            cleaned = f"{m.group(1)} {m.group(2)} {m.group(3)}"
        else:
            # Fallback for just MMM YYYY or generic stripping
            cleaned = re.sub(r'(\d{1,2})[.,]\s+', r'\1 ', cleaned)

        try:
            date = parser.parse(
                cleaned, dayfirst=True,
                default=_dt.datetime(2000, 1, 1)
            ).date()
            return date.strftime('%d/%m/%Y')
        except (ValueError, TypeError):
            return None

    def _rejoin_split_ocr_dates(self, ocr_lines):
        """
        OCR often reads '13 JAN 2022' as separate tokens across lines.
        Strategy: whenever we see a 'MMM YYYY' line, look back up to 5 lines
        for a standalone day number and prepend it.
        Also handles 'DD MMM' + year on next line.
        Returns a new list with extra joined candidates appended.
        """
        month_names = 'JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC'
        mon_yr_pat = re.compile(
            r'^(?:' + month_names + r')[A-Z]*\s+\d{4}$', re.IGNORECASE
        )
        day_pat = re.compile(r'^\d{1,2}$')
        dd_mon_pat = re.compile(r'^\d{1,2}\s+(?:' + month_names + r')[A-Z]*$', re.IGNORECASE)

        normalized = [self._normalize_ocr_line(l) for l in ocr_lines if l.strip()]
        rejoined = list(normalized)

        for i, line in enumerate(normalized):
            # Case 1: line is "MMM YYYY" — look back for a day digit
            if mon_yr_pat.match(line):
                for back in range(1, 6):
                    if i - back < 0:
                        break
                    prev = normalized[i - back].strip()
                    if day_pat.match(prev):
                        rejoined.append(f'{prev} {line}')
                        break

            # Case 2: line is "DD MMM" — look forward for year
            if dd_mon_pat.match(line):
                if i + 1 < len(normalized) and re.fullmatch(r'\d{4}', normalized[i + 1].strip()):
                    rejoined.append(f'{line} {normalized[i + 1].strip()}')

            # Case 3: consecutive lines are day then "MMM YYYY"
            if i + 1 < len(normalized):
                nxt = normalized[i + 1].strip()
                if day_pat.match(line.strip()) and mon_yr_pat.match(nxt):
                    rejoined.append(f'{line.strip()} {nxt}')

        return rejoined

    def clean(self, string):
        return ''.join(char for char in string if char.isalnum()).upper()

    def get_country_name(self, country_code):
        for country in self.country_codes:
            if country['code'] == country_code:
                return country['name']
        return country_code

    def find_authority(self, ocr_text):
        keywords = ['ISSUING AUTHORITY', 'ISSUED BY', 'AUTHORITY', 'ISSUING OFFICE', 'PLACE OF ISSUE']
        for line in ocr_text:
            for keyword in keywords:
                if keyword in line.upper():
                    authority = line.upper().split(keyword)[-1].strip()
                    return authority
        return 'Not Found'

    def _get_all_date_patterns(self):
        d = r'[0-9OISlB]'
        months = r'JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC'
        return [
            # DD-MM-YYYY / DD/MM/YY / DD.MM.YYYY
            f'{d}{{2}}[-/.]{d}{{2}}[-/.]{d}{{2,4}}',
            # YYYY-MM-DD
            f'{d}{{4}}[-/.]{d}{{2}}[-/.]{d}{{2}}',
            # DD MM YYYY (numeric with spaces)
            f'{d}{{1,2}} {d}{{1,2}} {d}{{2,4}}',
            # DD MMM YYYY (with potential noise between DD and MMM) e.g., 05 SA/MAR 2028 or 03.AUG 1991
            f'{d}{{1,2}}[.,\\sA-Z/]+(?:{months})[A-Z]*[.,\\s/]+{d}{{2,4}}',
            # MMM DD YYYY (with noise)
            f'\\b(?:{months})\\b[.,\\sA-Z/]+{d}{{1,2}}[.,\\sA-Z/]+{d}{{2,4}}',
            # MMM YYYY (fallback, e.g. if day is missing)
            f'\\b(?:{months})\\b\\s+{d}{{2,4}}',
        ]

    def _collect_all_dates(self, ocr_text):
        """
        Collect dates, preferring matches that have an explicit day component.
        For each (month, year) pair, keep only the entry with the most specific day.
        """
        mon_only_pat = re.compile(
            r'^(?:JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)[A-Z]*\s+\d{4}$', re.IGNORECASE
        )
        # Create a single joined string so regex '\s+' can bridge dates split across newlines
        full_text_blob = " ".join(ocr_text)
        search_lines = list(ocr_text) + [full_text_blob]
        
        # best[(month, year)] = (day, full_date_str)
        best = {}

        for line in search_lines:
            for pattern in self._get_all_date_patterns():
                for date_match in re.findall(pattern, line, re.IGNORECASE):
                    parsed = self._parse_ocr_date(date_match)
                    if not parsed:
                        continue
                    dd, mm, yyyy = parsed.split('/')
                    key = (mm, yyyy)
                    is_month_only = bool(mon_only_pat.match(date_match.strip()))
                    day_val = int(dd)
                    # Prefer: actual day > defaulted day (01 from MMM YYYY)
                    prev_day, _ = best.get(key, (0, None))
                    if not is_month_only and day_val > 1:
                        # Strong candidate — has explicit day
                        if prev_day <= 1:
                            best[key] = (day_val, parsed)
                    elif day_val > prev_day:
                        best[key] = (day_val, parsed)

        return {v for _, v in best.values()}

    def find_issuing_date(self, ocr_text, dob_str=None, expiry_str=None):
        """
        Smart Date Recognition: Picks the best 'Issue Date' by scoring candidates.
        Scores based on: Proximity to keywords, Plausibility (between DOB and Expiry).
        """
        candidates = self._collect_all_dates(ocr_text)
        if not candidates:
            return 'Not Found'

        # Parse comparison dates if available
        dob = None
        if dob_str and dob_str != 'Not Found':
            try:
                dob = parser.parse(dob_str, dayfirst=True).date()
            except Exception:
                pass

        expiry = None
        if expiry_str and expiry_str != 'Not Found':
            try:
                expiry = parser.parse(expiry_str, dayfirst=True).date()
            except Exception:
                pass

        scored_candidates = []
        issue_keywords = ['ISSUE', 'ISSUANCE', 'ISSUED', 'DATE OF ISSUE', 'DATE OF ISSUANCE']
        
        # We also look at the raw OCR lines for keyword proximity
        ocr_blob = " ".join(ocr_text).upper()

        for cand_str in candidates:
            score = 0
            try:
                cand_date = parser.parse(cand_str, dayfirst=True).date()
            except Exception:
                continue

            # 1. Keyword Proximity Score
            # Check if the year strongly appears in the same line as an issue keyword
            for line in ocr_text:
                if str(cand_date.year) in line:
                    if any(kw in line.upper() for kw in issue_keywords):
                        score += 10
                        break

            # 2. Strong Plausibility (Delta Years)
            if expiry:
                delta_years = expiry.year - cand_date.year
                if delta_years in [10, 5]:
                    score += 50  # Issue dates are typically 10 or 5 years before expiry
                elif delta_years > 0:
                    score += 10  # Valid past date relative to expiry
                elif delta_years < 0:
                    score -= 100 # Issue date cannot be AFTER expiry
            
            # General bounds
            if dob and cand_date > dob:
                score += 5
            if expiry and cand_date < expiry:
                score += 5

            # 3. Penalty for exact matches to DOB/Expiry (to avoid picking them)
            if dob_str == cand_str or expiry_str == cand_str:
                score -= 100
            if dob and cand_date == dob:
                score -= 100
            if expiry and cand_date == expiry:
                score -= 100

            scored_candidates.append((score, cand_str))

        if not scored_candidates:
            return 'Not Found'

        # Sort by score (descending) and then date (ascending)
        scored_candidates.sort(key=lambda x: (-x[0], x[1]))
        return scored_candidates[0][1]

    def print_data(self, data):
        for key, value in data.items():
            print(f'{key.replace("_", " ").capitalize()}\t:\t{value}')

    def _normalize_ocr_line(self, line):
        cleaned = re.sub(r'[\t\r\n]+', ' ', line)
        cleaned = re.sub(r'\s{2,}', ' ', cleaned)
        return cleaned.strip()

    def _normalize_mrz_name_token(self, token):
        value = (token or '').strip().upper()
        value = value.replace('«', '<').replace('|', '<')
        value = re.sub(r'<+', ' ', value)
        # Keep only letters/spaces for names.
        value = re.sub(r'[^A-Z ]+', ' ', value)
        value = re.sub(r'\s{2,}', ' ', value).strip()
        # OCR often reads MRZ filler "<" as repeated letters (e.g., KKKKKKKK).
        # Remove suspicious repeated trailing noise.
        value = re.sub(r'\b([A-Z])\1{2,}\b$', '', value).strip()
        # Remove trailing single-letter garbage tokens often produced by OCR filler
        # (e.g. "YAOZU K K" -> "YAOZU"). Keep initials only when the whole name is initials.
        tokens = [t for t in value.split(' ') if t]
        if any(len(t) > 1 for t in tokens):
            while tokens and len(tokens[-1]) == 1:
                tokens.pop()
        value = ' '.join(tokens).strip()
        return value

    def _mrz_name_quality(self, given, surname):
        full = f"{surname} {given}".strip()
        if not full:
            return -1
        letters = len(re.findall(r'[A-Z]', full))
        spaces = full.count(' ')
        penalties = 0
        if re.search(r'([A-Z])\1{3,}', full):
            penalties += 8
        if letters < 3:
            penalties += 5
        return letters + spaces - penalties

    def _parse_name_from_mrz_line(self, line_a):
        """
        Parse surname/given names from MRZ line 1 with tolerance for OCR noise.
        Handles formats like P<CHNSURNAME<<GIVEN<NAMES and common '<' misreads.
        """
        if not line_a:
            return '', ''

        raw = re.sub(r'\s+', '', line_a).upper()
        # Common OCR confusion in MRZ: '<' read as '«' or occasionally '|'
        raw = raw.replace('«', '<').replace('|', '<')

        # Try to split standard "....<<...." first.
        if '<<' in raw:
            left, right = raw.split('<<', 1)
        else:
            # Weak fallback: no clear separator means we cannot reliably infer both parts.
            return '', ''

        # Strip document prefix/country code, e.g. P<CHN or POCHN (OCR '<' -> 'O')
        if re.match(r'^P[<O][A-Z]{3}', left):
            left = left[5:]
        elif left.startswith('P<') and len(left) > 5:
            left = left[5:]

        surname = self._normalize_mrz_name_token(left)
        given = self._normalize_mrz_name_token(right)
        return given, surname

    def _extract_visual_latin_name(self, ocr_lines):
        """
        Extract name from visual zone lines like:
        'YANG, YAOZU' or 'TANG, FANGFANG'.
        Returns (given, surname) or ('', '').
        """
        stopwords = {
            'PEOPLE', 'REPUBLIC', 'CHINA', 'PASSPORT', 'MINISTRY', 'FOREIGN',
            'AFFAIRS', 'REQUESTS', 'AUTHORITIES', 'ALLOW', 'BEARER', 'ASSISTANCE',
            'DATE', 'BIRTH', 'ISSUE', 'EXPIRY', 'NATIONALITY', 'AUTHORITY'
        }
        for line in ocr_lines:
            n = self._normalize_ocr_line(line).upper()
            # Normalize separators and keep only useful symbols for matching.
            n = n.replace('，', ',')
            m = re.search(r'\b([A-Z]{2,})\s*,\s*([A-Z][A-Z ]{1,})\b', n)
            if not m:
                continue
            surname = self._normalize_mrz_name_token(m.group(1))
            given = self._normalize_mrz_name_token(m.group(2))

            if not surname or not given:
                continue

            surname_tokens = surname.split()
            given_tokens = given.split()
            all_tokens = surname_tokens + given_tokens

            # Safety filters to avoid selecting random paragraph text.
            if len(surname_tokens) != 1:
                continue
            if len(given_tokens) > 2:
                continue
            if any(len(t) == 1 for t in all_tokens):
                continue
            if any(t in stopwords for t in all_tokens):
                continue
            if len(''.join(all_tokens)) > 24:
                continue

            return given, surname
        return '', ''

    def _extract_given_by_surname_from_visual(self, ocr_lines, surname):
        """
        Find a visual-zone Latin line containing the detected surname and return
        the token immediately after it as given name.
        Example: "YANG, YAOZU" -> "YAOZU"
        """
        s = self._normalize_mrz_name_token(surname)
        if not s:
            return ''
        for line in ocr_lines:
            n = self._normalize_ocr_line(line).upper()
            # Keep letters and common separators, then split to tokens.
            n = re.sub(r'[^A-Z,.\- ]+', ' ', n)
            n = re.sub(r'[,.\-]+', ' ', n)
            tokens = [t for t in n.split() if t]
            if not tokens:
                continue
            for i, tok in enumerate(tokens[:-1]):
                if tok == s:
                    candidate = self._normalize_mrz_name_token(tokens[i + 1])
                    # Keep only realistic given-name token lengths.
                    if 2 <= len(candidate) <= 14:
                        return candidate
        return ''

    def _extract_labeled_fields(self, ocr_lines, label_map):
        normalized_lines = [self._normalize_ocr_line(line) for line in ocr_lines if line.strip()]
        results = {field: 'Not Found' for field in label_map.keys()}

        for i, line in enumerate(normalized_lines):
            upper_line = line.upper()

            for field, labels in label_map.items():
                if results[field] != 'Not Found':
                    continue

                for label in labels:
                    label_upper = label.upper()

                    if label_upper in upper_line:
                        # Try to split on ":" first, then on label occurrence
                        if ':' in line:
                            value = line.split(':', 1)[1].strip()
                        else:
                            parts = re.split(re.escape(label), line, flags=re.IGNORECASE, maxsplit=1)
                            value = parts[1].strip() if len(parts) > 1 else ''

                        # If value is empty, try next line as fallback
                        if not value and i + 1 < len(normalized_lines):
                            value = normalized_lines[i + 1].strip()

                        value = value.lstrip('.').strip()
                        results[field] = value if value else 'Not Found'
                        break

        return results

    def _extract_date_for_label(self, ocr_lines, label):
        normalized_lines = [self._normalize_ocr_line(line) for line in ocr_lines if line.strip()]
        label_upper = label.upper()
        date_patterns = self._get_all_date_patterns()

        for i, line in enumerate(normalized_lines):
            if label_upper in line.upper():
                # Proximity Check 1: Same line
                for pattern in date_patterns:
                    date_matches = re.findall(pattern, line, re.IGNORECASE)
                    for date_match in date_matches:
                        parsed_date = self._parse_ocr_date(date_match)
                        if parsed_date:
                            return parsed_date

                # Proximity Check 2: Immediately following line
                if i + 1 < len(normalized_lines):
                    next_line = normalized_lines[i + 1]
                    # To avoid picking a date from a totally different section,
                    # we ensure the next line is "close" (doesn't look like a new field header)
                    if not any(header in next_line.upper() for header in ['NAME', 'PASSPORT', 'AUTHORITY', 'BIRTH']):
                        for pattern in date_patterns:
                            date_matches = re.findall(pattern, next_line, re.IGNORECASE)
                            for date_match in date_matches:
                                parsed_date = self._parse_ocr_date(date_match)
                                if parsed_date:
                                    return parsed_date

        return 'Not Found'

    def _extract_date_for_labels(self, ocr_lines, labels):
        for label in labels:
            found = self._extract_date_for_label(ocr_lines, label)
            if found != 'Not Found':
                return found
        return 'Not Found'

    def _extract_date_near_keywords(self, ocr_lines, keywords):
        normalized_lines = [self._normalize_ocr_line(line) for line in ocr_lines if line.strip()]
        keywords_upper = [k.upper() for k in keywords]
        date_patterns = self._get_all_date_patterns()
        month_pat = r'(?:JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)[A-Z]*'

        for i, line in enumerate(normalized_lines):
            upper_line = line.upper()
            if any(k in upper_line for k in keywords_upper):
                for pattern in date_patterns:
                    for m in re.finditer(pattern, line, re.IGNORECASE):
                        parsed = self._parse_ocr_date(m.group())
                        if parsed:
                            return parsed

                for offset in (1, 2):
                    if i + offset < len(normalized_lines):
                        next_line = normalized_lines[i + offset]
                        for pattern in date_patterns:
                            for m in re.finditer(pattern, next_line, re.IGNORECASE):
                                parsed = self._parse_ocr_date(m.group())
                                if parsed:
                                    return parsed

                block = ' '.join(normalized_lines[i:i + 6])
                for pattern in date_patterns:
                    for m in re.finditer(pattern, block, re.IGNORECASE):
                        parsed = self._parse_ocr_date(m.group())
                        if parsed:
                            return parsed
                pat = r'(\d{1,2})\s+' + month_pat + r'(?:\s+' + month_pat + r')?\s+(\d{2,4})'
                for m in re.finditer(pat, block, re.IGNORECASE):
                    parts = m.group(0).split()
                    s = f"{parts[0]} {parts[1]} {parts[-1]}"
                    parsed = self._parse_ocr_date(s)
                    if parsed:
                        return parsed

        return 'Not Found'

    def _split_date_components(self, date_string):
        if not date_string or date_string == 'Not Found':
            return 'Not Found', 'Not Found', 'Not Found'

        try:
            parsed = parser.parse(date_string, dayfirst=True).date()
        except (ValueError, TypeError):
            return 'Not Found', 'Not Found', 'Not Found'

        day = f'{parsed.day:02d}'
        month = f'{parsed.month:02d}'
        year = f'{parsed.year:04d}'
        return day, month, year

    def _ocr_lines(self, img, engine='both'):
        """Return OCR lines from the requested engine: 'easyocr', 'tesseract', or 'both'."""
        if engine == 'easyocr':
            return self._easyocr_lines(img)
        elif engine == 'tesseract':
            if not self._tesseract_available:
                # Graceful fallback when Tesseract is missing.
                return self._easyocr_lines(img)
            return self._tesseract_lines(img)
        if engine == 'both' and not self._tesseract_available:
            return self._easyocr_lines(img)
        return self._dual_ocr_lines(img)

    def get_data(self, img_name, debug=False, ocr_engine='both'):
        user_info = {}
        with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmpfile:
            tmpfile_path = tmpfile.name

        try:
            try:
                mrz = read_mrz(img_name, save_roi=True)
            except (pytesseract.pytesseract.TesseractNotFoundError, FileNotFoundError, OSError):
                # passporteye can invoke Tesseract internally for MRZ extraction.
                # If unavailable, continue without MRZ instead of crashing.
                mrz = None
                self._tesseract_available = False
            if mrz:
                mpimg.imsave(tmpfile_path, mrz.aux['roi'], cmap='gray')
                img = cv2.imread(tmpfile_path)
                img = cv2.resize(img, (1110, 140))
                allowlist = st.ascii_letters + st.digits + '< '
                code = self.reader.readtext(img, paragraph=False, detail=0, allowlist=allowlist)

                if len(code) < 2:
                    return print(f'Error: Insufficient OCR results for image {img_name}.')

                # Strip spaces from MRZ lines — OCR sometimes inserts spurious spaces
                a = re.sub(r'\s+', '', code[0]).upper()
                b = re.sub(r'\s+', '', code[1]).upper()

                if len(a) < 44:
                    a = a + '<' * (44 - len(a))
                if len(b) < 44:
                    b = b + '<' * (44 - len(b))

                name, surname = self._parse_name_from_mrz_line(a)

                # Prefer passporteye's highly accurate built-in MRZ parser over manual EasyOCR string splitting
                mrz_data = mrz.to_dict()
                if mrz_data.get('names') and mrz_data.get('surname'):
                    parsed_given = self._normalize_mrz_name_token(mrz_data.get('names'))
                    parsed_surname = self._normalize_mrz_name_token(mrz_data.get('surname'))
                    # Keep whichever source looks cleaner: passporteye parsed fields vs OCR MRZ line parse.
                    if self._mrz_name_quality(parsed_given, parsed_surname) >= self._mrz_name_quality(name, surname):
                        name, surname = parsed_given, parsed_surname
              
                full_img = cv2.imread(img_name)
                ocr_results = self._ocr_lines(full_img, engine=ocr_engine)
                ocr_extended = self._rejoin_split_ocr_dates(ocr_results)

                # Visual-zone fallback can correct MRZ OCR noise (e.g. trailing "K K").
                v_given, v_surname = self._extract_visual_latin_name(ocr_results)
                if self._mrz_name_quality(v_given, v_surname) > self._mrz_name_quality(name, surname):
                    name, surname = v_given, v_surname

                # Final targeted correction: if visual line has a clean token right after
                # the chosen surname, prefer it over noisy MRZ given names (e.g. YAROZU -> YAOZU).
                given_from_surname = self._extract_given_by_surname_from_visual(ocr_results, surname)
                if given_from_surname and (
                    not name or
                    name == 'Not Found' or
                    len(given_from_surname) < len(name) + 2
                ):
                    name = given_from_surname

                user_info['Given Names'] = name if name else 'Not Found'
                user_info['Surname'] = surname if surname else 'Not Found'
                full_name = ' '.join(
                    [p for p in [user_info['Surname'], user_info['Given Names']] if p != 'Not Found']
                )
                user_info['Name'] = full_name if full_name else 'Not Found'
                dob = self.parse_birth_date(b[13:19])
                expiry = self.parse_date(b[21:27])
                user_info['Date of Birth'] = dob

                issue_date = self._extract_date_for_labels(
                    ocr_extended,
                    ['DATE OF ISSUE', 'DATE OF ISSUANCE', 'ISSUED ON', 'ISSUED DATE', 'ISSUE DATE',
                     'Date of Issue', 'Dale pf issue', 'Dale of issue']
                )
                if issue_date == 'Not Found':
                    issue_date = self._extract_date_near_keywords(
                        ocr_extended,
                        ['DATE OF ISSUE', 'DATE OF ISTUE', 'DALE PF ISSUE', 'DALE OF ISSUE', 'ISSUANCE']
                    )
                if issue_date == 'Not Found':
                    all_tess = self._rejoin_split_ocr_dates(ocr_results)
                    issue_date = self.find_issuing_date(all_tess, dob_str=dob, expiry_str=expiry)
                if debug:
                    print('DEBUG full OCR lines:')
                    for line in ocr_results:
                        print(f'  {line!r}')
                    all_dates = self._collect_all_dates(ocr_results)
                    print('DEBUG all dates found:', all_dates)
                user_info['Date of Issue'] = issue_date
                user_info['Date of Expiry'] = expiry
                user_info['Authority'] = self.find_authority(ocr_results)
                user_info['Nationality'] = self.get_country_name(self.clean(b[10:13]))
                user_info['Passport Type'] = self.clean(a[0:2])
                user_info['Passport Number'] = self.clean(b[0:9])
                # Gender: from MRZ only (passporteye's parsed sex, or MRZ line 2 position 20)
                sex = (getattr(mrz, 'sex', '') or '').upper()
                if sex in ('M', 'F'):
                    user_info['Gender'] = 'Male' if sex == 'M' else 'Female'
                elif len(b) > 20 and b[20] in ('M', 'F'):
                    user_info['Gender'] = 'Male' if b[20] == 'M' else 'Female'
                else:
                    user_info['Gender'] = 'Not Found'

                # Fallback for Gender if MRZ fails
                if user_info['Gender'] == 'Not Found':
                    gender_results = self._extract_labeled_fields(ocr_results, {'Gender': ['GENDER']})
                    raw_g = gender_results.get('Gender', 'Not Found')
                    if raw_g != 'Not Found':
                        user_info['Gender'] = 'Male' if 'M' in raw_g.upper() else ('Female' if 'F' in raw_g.upper() else 'Not Found')

                # Combined Card Detection: check if this image also contains a Foreign Employment Card
                # Use a more flexible search for the card header
                combined_text = ' '.join(ocr_results).upper()
                card_keywords = ['FOREIGN EMPLOYMENT CARD', 'EMPLOYMENT CARD', 'FWCMS']
                is_card_detected = any(kw in combined_text for kw in card_keywords)
                
                if debug:
                    print(f"DEBUG: Combined card detected: {is_card_detected}")
                
                if is_card_detected:
                    card_data = self.get_foreign_employment_card_data(img_name, ocr_engine=ocr_engine)
                    if debug:
                        print(f"DEBUG: Extracted card data: {card_data}")
                    user_info.update(card_data)


            else:
                print(f'Machine cannot read MRZ from image {img_name}.')

        finally:
            if os.path.exists(tmpfile_path):
                os.remove(tmpfile_path)

        return user_info

    def get_foreign_employment_card_data(self, img_name, ocr_engine='both'):
        full_img = cv2.imread(img_name)
        if full_img is None:
            print(f'Image not found or unreadable: {img_name}')
            return {}

        ocr_results = self._ocr_lines(full_img, engine=ocr_engine)

        label_map = {
            'Company': ['COMPANY', 'COMPANY NAME', 'ENTERPRISE NAME'],
            'Position': ['POSITION', 'POSTION', 'POSITOIN'],
            'Card Number': ['CARD NUMBER', 'CARD NO', 'ID NO', 'ID NO.', 'I0 NO', 'I0 NO.'],
            'DC1': ['DC1', 'EXPIRED DATE', 'EXPIRY DATE', 'DATE OF EXPIRY'],
            'DC2': ['DC2'],
            'DC3': ['DC3'],
            'Company Card': ['COMPANY CARD'],
            'Position Card': ['POSITION CARD', 'POSITOIN CARD'],
            'Phone': ['PHONE', 'TEL', 'MOBILE'],
            'D01': ['D01'],
            'D02': ['D02'],
            'D03': ['D03'],
            'Name 02': ['NAME 02', 'NAME02'],
        }

        results = self._extract_labeled_fields(ocr_results, label_map)

        # If DC1/DC2/DC3 represent Expired Date (day/month/year), derive them.
        expired_date = self._extract_date_for_label(ocr_results, 'Expired Date')
        if expired_date == 'Not Found':
            expired_date = results.get('DC1', 'Not Found')
        day, month, year = self._split_date_components(expired_date)

        if day != 'Not Found':
            results['DC1'] = day
        if month != 'Not Found':
            results['DC2'] = month
        if year != 'Not Found':
            results['DC3'] = year

        # Map Company Card and Position Card from enterprise name and position.
        if results.get('Company Card') == 'Not Found':
            results['Company Card'] = results.get('Company', 'Not Found')
        if results.get('Position Card') == 'Not Found':
            results['Position Card'] = results.get('Position', 'Not Found')

        return results

    def _split_name(self, full_name):
        if not full_name or full_name == 'Not Found':
            return 'Not Found', 'Not Found'
        parts = [p for p in full_name.strip().split(' ') if p]
        if not parts:
            return 'Not Found', 'Not Found'
        surname = parts[-1]
        given = ' '.join(parts[:-1]) if len(parts) > 1 else 'Not Found'
        return surname, given

    def _build_combined(self, passport_data, card_data):
        surname = passport_data.get('Surname', 'Not Found')
        given = passport_data.get('Given Names', 'Not Found')
        if surname == 'Not Found' and given == 'Not Found':
            surname, given = self._split_name(passport_data.get('Name', 'Not Found'))
        bd1, bd2, bd3 = self._split_date_components(passport_data.get('Date of Birth', 'Not Found'))
        iss1, iss2, iss3 = self._split_date_components(passport_data.get('Date of Issue', 'Not Found'))
        ed1, ed2, ed3 = self._split_date_components(passport_data.get('Date of Expiry', 'Not Found'))
        # Keep NAME 02 consistent with form layout: SURNAME first, then GSURNAME.
        full_name_parts = []
        if surname and surname != 'Not Found':
            full_name_parts.append(surname)
        if given and given != 'Not Found':
            full_name_parts.append(given)
        full_name = ' '.join(full_name_parts) if full_name_parts else 'Not Found'
        return {
            'SURNAME':       surname,
            'GSURNAME':      given,
            'BD1': bd1, 'BD2': bd2, 'BD3': bd3,
            'NASTIONALTY':   passport_data.get('Nationality', 'Not Found'),
            'PASSPORT':      passport_data.get('Passport Number', 'Not Found'),
            'ISS1': iss1, 'ISS2': iss2, 'ISS3': iss3,
            'ED1': ed1, 'ED2': ed2, 'ED3': ed3,
            'CARD NUMBER':   card_data.get('Card Number', 'Not Found'),
            'DC1':           card_data.get('DC1', 'Not Found'),
            'DC2':           card_data.get('DC2', 'Not Found'),
            'DC3':           card_data.get('DC3', 'Not Found'),
            'COMPANY CARD':  card_data.get('Company Card', 'Not Found'),
            'POSITOIN CARD': card_data.get('Position Card', 'Not Found'),
            'NAME 02':       full_name,
            'GENDER_RAW':    passport_data.get('Gender', ''),
            'Gender':        self._normalize_gender(passport_data.get('Gender', '')),
        }

    def _normalize_gender(self, raw: str) -> str:
        """Return M or F from MRZ Gender, or empty string."""
        g = (raw or '').upper().strip()
        # Check F first: "Female" contains 'M' so would wrongly match Male
        if g in ('F', 'FEMALE') or g.startswith('F'):
            return 'F'
        if g in ('M', 'MALE') or g.startswith('M'):
            return 'M'
        return ''

    def get_passport_and_card_data(self, passport_img_name, card_img_name,
                                   debug=False, ocr_engine='both'):
        passport_data = self.get_data(passport_img_name, debug=debug, ocr_engine=ocr_engine)
        card_data = self.get_foreign_employment_card_data(card_img_name, ocr_engine=ocr_engine)
        return self._build_combined(passport_data, card_data)

    # Column headers for PASSPORT-FORM template (skip COMPANY, POSTION, PHONE)
    EXCEL_HEADERS = [
        'SURNAME', 'GSURNAME',
        'BD1', 'BD2', 'BD3',
        'NASTIONALTY', 'PASSPORT',
        'ISS1', 'ISS2', 'ISS3',
        'ED1', 'ED2', 'ED3',
        'COMPANY', 'POSTION',
        'CARD NUMBER',
        'DC1', 'DC2', 'DC3',
        'COMPANY CARD', 'POSITOIN CARD',
        'PHONE',
        'D01', 'D02', 'D03',
        'NAME 02',
        'M', 'F',
        '12M', '6M', '3M', '1M', 'M_VAL',
    ]

    def save_to_excel(self, data, xlsx_path, validity_period=None, gender_override=None):
        """
        Append a row into xlsx_path. Creates the file with headers if it doesn't exist.
        """
        if os.path.exists(xlsx_path):
            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'PASSPORT FORM'
            for col, header in enumerate(self.EXCEL_HEADERS, start=1):
                # Write 'M' to Excel for both Gender M and Validity M
                val = 'M' if header == 'M_VAL' else header
                ws.cell(row=1, column=col, value=val)

        # Build header → column index map from row 1
        # To handle duplicate 'M', we use fixed indices for Gender M vs Validity M if possible,
        # or we just rely on the order in EXCEL_HEADERS.
        header_map = {}
        for col, h in enumerate(self.EXCEL_HEADERS, start=1):
            header_map[h.upper()] = col

        # Find next empty data row
        next_row = 2
        for row in range(2, ws.max_row + 2):
            if all(ws.cell(row=row, column=c).value in (None, '') for c in range(1, ws.max_column + 1)):
                next_row = row
                break

        # Write each field into the correct column (skip internal helpers)
        skip_fields = {'GENDER_RAW'}
        for field, value in data.items():
            if field in skip_fields:
                continue
            col = header_map.get(field.strip().upper())
            if col and value and value != 'Not Found':
                ws.cell(row=next_row, column=col, value=value)

        # D01, D02, D03 = date current (today)
        import datetime
        today = datetime.date.today()
        for hdr, val in [('D01', str(today.day).zfill(2)), ('D02', str(today.month).zfill(2)), ('D03', str(today.year))]:
            col = header_map.get(hdr)
            if col:
                ws.cell(row=next_row, column=col, value=val)

        # M / F tick from Gender (check F first: "Female" contains 'M')
        gender = (gender_override or data.get('GENDER_RAW') or '').upper().strip()
        if gender in ('F', 'FEMALE') or gender.startswith('F'):
            col = header_map.get('F')
            if col:
                ws.cell(row=next_row, column=col, value='X')
        elif gender in ('M', 'MALE') or gender.startswith('M'):
            col = header_map.get('M')
            if col:
                ws.cell(row=next_row, column=col, value='X')

        # Validity period tick (12M, 6M, 3M, 1M, M_VAL)
        v_period = str(validity_period).upper() if validity_period else ""
        if v_period == 'M':
            v_period = 'M_VAL'
            
        if v_period in ('12M', '6M', '3M', '1M', 'M_VAL'):
            col = header_map.get(v_period)
            if col:
                ws.cell(row=next_row, column=col, value='X')

        wb.save(xlsx_path)
        print(f'Saved row {next_row} to {xlsx_path}')

    def save_many_to_excel(self, data_list, xlsx_path, validity_period=None, gender_override=None, is_exported_list=None):
        if not data_list:
            return
        from openpyxl.styles import PatternFill
        grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        if os.path.exists(xlsx_path):
            import openpyxl
            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb.active
        else:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'PASSPORT FORM'
            for col, header in enumerate(self.EXCEL_HEADERS, start=1):
                # Write 'M' to Excel for both Gender M and Validity M
                val = 'M' if header == 'M_VAL' else header
                ws.cell(row=1, column=col, value=val)

        # Build header → column index map from row 1
        header_map = {}
        for col, h in enumerate(self.EXCEL_HEADERS, start=1):
            header_map[h.upper()] = col

        next_row = 2
        for row in range(2, ws.max_row + 2):
            if all(ws.cell(row=row, column=c).value in (None, '') for c in range(1, ws.max_column + 1)):
                next_row = row
                break

        import datetime
        today = datetime.date.today()
        skip_fields = {'GENDER_RAW'}

        for data in data_list:
            for field, value in data.items():
                if field in skip_fields:
                    continue
                col = header_map.get(field.strip().upper())
                if col and value and value != 'Not Found':
                    ws.cell(row=next_row, column=col, value=value)

            for hdr, val in [('D01', str(today.day).zfill(2)), ('D02', str(today.month).zfill(2)), ('D03', str(today.year))]:
                col = header_map.get(hdr)
                if col:
                    ws.cell(row=next_row, column=col, value=val)

            gender = (gender_override or data.get('GENDER_RAW') or '').upper().strip()
            if gender in ('F', 'FEMALE') or gender.startswith('F'):
                col = header_map.get('F')
                if col:
                    ws.cell(row=next_row, column=col, value='X')
            elif gender in ('M', 'MALE') or gender.startswith('M'):
                col = header_map.get('M')
                if col:
                    ws.cell(row=next_row, column=col, value='X')

            # Validity period tick (12M, 6M, 3M, 1M, M_VAL)
            v_period = str(validity_period).upper() if validity_period else ""
            if v_period == 'M':
                v_period = 'M_VAL'
                
            if v_period in ('12M', '6M', '3M', '1M', 'M_VAL'):
                col = header_map.get(v_period)
                if col:
                    ws.cell(row=next_row, column=col, value='X')
            
            is_exported = is_exported_list[data_list.index(data)] if is_exported_list and data_list.index(data) < len(is_exported_list) else False
            if is_exported:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=next_row, column=col).fill = grey_fill

            next_row += 1

        wb.save(xlsx_path)
        print(f'Saved {len(data_list)} rows to {xlsx_path}')

    def get_passport_and_card_data_all_engines(self, passport_img_name, card_img_name, debug=False):
        """
        Run extraction with EasyOCR, Tesseract, and Both engines separately.
        Returns a dict with keys 'easyocr', 'tesseract', 'both'.
        Also prints a side-by-side comparison table.
        """
        results = {}
        for engine in ('easyocr', 'tesseract', 'both'):
            passport_data = self.get_data(passport_img_name, debug=debug, ocr_engine=engine)
            card_data = self.get_foreign_employment_card_data(card_img_name, ocr_engine=engine)
            results[engine] = self._build_combined(passport_data, card_data)

        # Print comparison table
        fields = list(results['easyocr'].keys())
        col_w = 22
        header = f"{'FIELD':<18}  {'EASYOCR':<{col_w}}  {'TESSERACT':<{col_w}}  {'BOTH':<{col_w}}"
        print(header)
        print('-' * len(header))
        for field in fields:
            easy = results['easyocr'].get(field, '')
            tess = results['tesseract'].get(field, '')
            both = results['both'].get(field, '')
            print(f"{field:<18}  {easy:<{col_w}}  {tess:<{col_w}}  {both:<{col_w}}")

        return results

# Usage Example
if __name__ == "__main__":
    country_codes_file = 'data/country_codes.json'
    passport_img = 'images/pass_empoy.jpg'
    employee_card_img = 'images/pass_empoy.jpg'
    xlsx_path = 'PASSPORT-FORM.xlsx'

    extractor = PassportDataExtractor(country_codes_file)
    results = extractor.get_passport_and_card_data_all_engines(passport_img, employee_card_img)

    # Save the 'both' engine result to Excel
    extractor.save_to_excel(results['both'], xlsx_path)
