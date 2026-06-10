import os
import datetime
import openpyxl
from openpyxl.styles import PatternFill


class ExcelMixin:
    EXCEL_HEADERS = [
        'SURNAME', 'GSURNAME',
        'BD1', 'BD2', 'BD3',
        'NASTIONALTY', 'PASSPORT',
        'ISS1', 'ISS2', 'ISS3',
        'ED1', 'ED2', 'ED3',
        'COMPANY', 'POSTION',
        'CARD NUMBER',
        'DC1', 'DC2', 'DC3',
        'COMPANY CARD', 'POSITOIN CARD',
        'PHONE',
        'D01', 'D02', 'D03',
        'NAME 02',
        'M', 'F',
        '12M', '6M', '3M', '1M', 'M_VAL',
    ]

    def save_to_excel(self, data, xlsx_path, validity_period=None, gender_override=None):
        if os.path.exists(xlsx_path):
            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'PASSPORT FORM'
            for col, header in enumerate(self.EXCEL_HEADERS, start=1):
                val = 'M' if header == 'M_VAL' else header
                ws.cell(row=1, column=col, value=val)

        header_map = {h.upper(): col for col, h in enumerate(self.EXCEL_HEADERS, start=1)}

        next_row = 2
        for row in range(2, ws.max_row + 2):
            if all(ws.cell(row=row, column=c).value in (None, '') for c in range(1, ws.max_column + 1)):
                next_row = row
                break

        skip_fields = {'GENDER_RAW'}
        for field, value in data.items():
            if field in skip_fields:
                continue
            col = header_map.get(field.strip().upper())
            if col and value and value != 'Not Found':
                ws.cell(row=next_row, column=col, value=value)

        today = datetime.date.today()
        for hdr, val in [('D01', str(today.day).zfill(2)), ('D02', str(today.month).zfill(2)), ('D03', str(today.year))]:
            col = header_map.get(hdr)
            if col:
                ws.cell(row=next_row, column=col, value=val)

        gender = (gender_override or data.get('GENDER_RAW') or '').upper().strip()
        if gender in ('F', 'FEMALE') or gender.startswith('F'):
            col = header_map.get('F')
            if col:
                ws.cell(row=next_row, column=col, value='X')
        elif gender in ('M', 'MALE') or gender.startswith('M'):
            col = header_map.get('M')
            if col:
                ws.cell(row=next_row, column=col, value='X')

        v_period = str(validity_period).upper() if validity_period else ""
        if v_period == 'M':
            v_period = 'M_VAL'
        if v_period in ('12M', '6M', '3M', '1M', 'M_VAL'):
            col = header_map.get(v_period)
            if col:
                ws.cell(row=next_row, column=col, value='X')

        wb.save(xlsx_path)
        print(f'Saved row {next_row} to {xlsx_path}')

    def save_many_to_excel(self, data_list, xlsx_path, validity_period=None, gender_override=None, is_exported_list=None):
        if not data_list:
            return
        grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        if os.path.exists(xlsx_path):
            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'PASSPORT FORM'
            for col, header in enumerate(self.EXCEL_HEADERS, start=1):
                val = 'M' if header == 'M_VAL' else header
                ws.cell(row=1, column=col, value=val)

        header_map = {h.upper(): col for col, h in enumerate(self.EXCEL_HEADERS, start=1)}

        next_row = 2
        for row in range(2, ws.max_row + 2):
            if all(ws.cell(row=row, column=c).value in (None, '') for c in range(1, ws.max_column + 1)):
                next_row = row
                break

        today = datetime.date.today()
        skip_fields = {'GENDER_RAW'}

        for data in data_list:
            for field, value in data.items():
                if field in skip_fields:
                    continue
                col = header_map.get(field.strip().upper())
                if col and value and value != 'Not Found':
                    ws.cell(row=next_row, column=col, value=value)

            for hdr, val in [('D01', str(today.day).zfill(2)), ('D02', str(today.month).zfill(2)), ('D03', str(today.year))]:
                col = header_map.get(hdr)
                if col:
                    ws.cell(row=next_row, column=col, value=val)

            gender = (gender_override or data.get('GENDER_RAW') or '').upper().strip()
            if gender in ('F', 'FEMALE') or gender.startswith('F'):
                col = header_map.get('F')
                if col:
                    ws.cell(row=next_row, column=col, value='X')
            elif gender in ('M', 'MALE') or gender.startswith('M'):
                col = header_map.get('M')
                if col:
                    ws.cell(row=next_row, column=col, value='X')

            v_period = str(validity_period).upper() if validity_period else ""
            if v_period == 'M':
                v_period = 'M_VAL'
            if v_period in ('12M', '6M', '3M', '1M', 'M_VAL'):
                col = header_map.get(v_period)
                if col:
                    ws.cell(row=next_row, column=col, value='X')

            is_exported = is_exported_list[data_list.index(data)] if is_exported_list and data_list.index(data) < len(is_exported_list) else False
            if is_exported:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=next_row, column=col).fill = grey_fill

            next_row += 1

        wb.save(xlsx_path)
        print(f'Saved {len(data_list)} rows to {xlsx_path}')
